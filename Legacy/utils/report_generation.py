# report_generation.py
# Functions to generate reports

import numpy as np
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def generate_old_events_properties_report(stale_events_df, single_day_events_df, stale_properties_df, 
                                          single_day_properties_df, stale_user_props_df, 
                                          single_day_user_props_df, output_dir):
    """
    Generates a report for old (stale) and single-day events, event properties, and user properties.

    Parameters:
    - stale_events_df (pd.DataFrame): Events not seen in over a year.
    - single_day_events_df (pd.DataFrame): Events seen only for one day.
    - stale_properties_df (pd.DataFrame): Event properties not seen in over a year.
    - single_day_properties_df (pd.DataFrame): Event properties seen only for one day.
    - stale_user_props_df (pd.DataFrame): User properties not seen in over a year.
    - single_day_user_props_df (pd.DataFrame): User properties seen only for one day.
    - output_dir (str): Directory to save the report.
    """
    report_file = os.path.join(output_dir, "stale_and_single_day_events_properties_report.xlsx")
    
    with pd.ExcelWriter(report_file, engine='openpyxl') as writer:
        any_written = False

        if not stale_events_df.empty:
            stale_events_df.to_excel(writer, sheet_name="Stale Events", index=False)
            any_written = True
        if not single_day_events_df.empty:
            single_day_events_df.to_excel(writer, sheet_name="Single-Day Events", index=False)
            any_written = True
        if not stale_properties_df.empty:
            stale_properties_df.to_excel(writer, sheet_name="Stale Properties", index=False)
            any_written = True
        if not single_day_properties_df.empty:
            single_day_properties_df.to_excel(writer, sheet_name="Single-Day Properties", index=False)
            any_written = True
        if not stale_user_props_df.empty:
            stale_user_props_df.to_excel(writer, sheet_name="Stale User Properties", index=False)
            any_written = True
        if not single_day_user_props_df.empty:
            single_day_user_props_df.to_excel(writer, sheet_name="Single-Day User Properties", index=False)
            any_written = True

        if not any_written:
            pd.DataFrame({"Message": ["No stale or single-day events or properties found."]}) \
                .to_excel(writer, sheet_name="No Data", index=False)


    print(f"Old events, properties, and user properties report saved to {report_file}")


def generate_syntax_report(syntax_results, output_dir):
    """
    Generates a report on naming syntax for events, event properties, and user properties.

    Parameters:
    - syntax_results (dict): Dictionary containing categorized DataFrames.
    - output_dir (str): Directory to save the report.

    Returns:
    - pd.DataFrame: Summary DataFrame with syntax type counts for events, event properties, and user properties.
    """
    report_file = os.path.join(output_dir, "naming_syntax_report.xlsx")
    summary_data = []

    # Generate counts for the summary
    for category, df in syntax_results.items():
        if not df.empty:
            syntax_counts = df["Syntax Category"].value_counts().reset_index()
            syntax_counts.columns = ["Syntax Type", "Count"]
            syntax_counts["Data Type"] = category.replace("_", " ").title()
            summary_data.append(syntax_counts)

    with pd.ExcelWriter(report_file, engine="openpyxl") as writer:
        any_written = False

        # Write summary if any data exists
        if summary_data:
            summary_df = pd.concat(summary_data)
            pivot_summary = summary_df.pivot_table(
                index="Syntax Type",
                columns="Data Type",
                values="Count",
                fill_value=0
            )
            pivot_summary.to_excel(writer, sheet_name="Summary")
            any_written = True
        else:
            pivot_summary = pd.DataFrame()  # still return empty summary if nothing matched

        # Write each category's raw data if available
        for category, df in syntax_results.items():
            if not df.empty:
                df.to_excel(writer, sheet_name=category.replace("_", " ").title(), index=False)
                any_written = True

        # Fallback sheet if no data was written
        if not any_written:
            pd.DataFrame({"Message": ["No naming syntax patterns found in project data."]}) \
                .to_excel(writer, sheet_name="No Data", index=False)

    print(f"✅ Naming Syntax Report saved to {report_file}")
    return pivot_summary.reset_index() if not pivot_summary.empty else pd.DataFrame()



def generate_unused_events_report(top_unused_events, bottom_unused_events, output_dir, lookback_window):
    """
    Generates a report for unused events based on the selected look-back window.

    Parameters:
    - top_unused_events (pd.DataFrame): DataFrame of top 10 high-volume unused events.
    - bottom_unused_events (pd.DataFrame): DataFrame of bottom 10 low-volume unused events.
    - output_dir (str): Directory to save the report.
    - lookback_window (int): The selected look-back window in days.
    """
    report_file = os.path.join(output_dir, f"unused_events_report_{lookback_window}d.xlsx")

    with pd.ExcelWriter(report_file, engine="openpyxl") as writer:
        any_written = False

        if not top_unused_events.empty:
            top_unused_events.to_excel(writer, sheet_name="Top 10 Unused Events", index=False)
            any_written = True

        if not bottom_unused_events.empty:
            bottom_unused_events.to_excel(writer, sheet_name="Bottom 10 Unused Events", index=False)
            any_written = True

        if not any_written:
            pd.DataFrame({"Message": ["No unused events found for this project."]}) \
                .to_excel(writer, sheet_name="No Data", index=False)

    print(f"Unused events report saved to {report_file}")


def generate_missing_categories_descriptions_report(
    summary_df, 
    missing_event_categories, 
    missing_event_descriptions, 
    missing_event_prop_descriptions, 
    missing_user_prop_descriptions, 
    output_dir
):
    report_file = os.path.join(output_dir, "missing_categories_descriptions_report.xlsx")

    # Helper function to sanitize dataframes
    def sanitize_df(df, numeric_cols=None):
        if df.empty:
            return df
        df = df.copy()
        df.columns = [str(col).strip().replace('\n', ' ').replace('\r', '') for col in df.columns]
        for col in df.columns:
            df[col] = df[col].apply(lambda x: '' if pd.isna(x) else str(x).strip())
        if numeric_cols:
            for col in numeric_cols:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
        return df

    summary_df = sanitize_df(summary_df, numeric_cols=["Count"])
    missing_event_categories = sanitize_df(missing_event_categories)
    missing_event_descriptions = sanitize_df(missing_event_descriptions)
    missing_event_prop_descriptions = sanitize_df(missing_event_prop_descriptions)
    missing_user_prop_descriptions = sanitize_df(missing_user_prop_descriptions)

    wb = Workbook()

    def add_df_to_sheet(df, sheet_name):
        ws = wb.create_sheet(title=sheet_name)
        if df.empty:
            ws.append(["No Data"])
            return
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

    # Add sheets with SHORTER NAMES
    add_df_to_sheet(summary_df, "Summary")
    add_df_to_sheet(missing_event_categories, "Missing Event Categories")
    add_df_to_sheet(missing_event_descriptions, "Missing Event Descriptions")
    add_df_to_sheet(missing_event_prop_descriptions, "Event Prop Missing Descriptions")
    add_df_to_sheet(missing_user_prop_descriptions, "User Prop Missing Descriptions")

    # Remove default blank sheet created by openpyxl
    default_sheet = wb["Sheet"] if "Sheet" in wb.sheetnames else None
    if default_sheet:
        wb.remove(default_sheet)

    wb.save(report_file)

    print(f"Missing Categories & Descriptions Report saved to {report_file}")
    
def generate_duplicate_events_report(duplicate_events_df, output_dir, lookback_window):
    """
    Generates a report for duplicate events with identical volumes.

    Parameters:
    - duplicate_events_df (pd.DataFrame): DataFrame containing duplicate events.
    - output_dir (str): Directory to save the report.
    - lookback_window (int): The selected look-back window in days.
    """
    report_file = os.path.join(output_dir, f"duplicate_events_report_{lookback_window}d.xlsx")

    with pd.ExcelWriter(report_file, engine="openpyxl") as writer:
        if not duplicate_events_df.empty:
            duplicate_events_df.to_excel(writer, sheet_name="Duplicate Events", index=False)
        else:
            pd.DataFrame({"Message": ["No duplicate events found."]}) \
              .to_excel(writer, sheet_name="No Data", index=False)

    print(f"Duplicate Events Report saved to {report_file}")
    

def generate_user_property_misclassification_report(flagged_properties_df, output_dir):
    """
    Generates a report for event properties that should be categorized as user properties.

    Parameters:
    - flagged_properties_df (pd.DataFrame): DataFrame containing flagged event properties.
    - output_dir (str): Directory to save the report.
    """
    report_file = os.path.join(output_dir, "user_property_misclassification_report.xlsx")

    with pd.ExcelWriter(report_file, engine='openpyxl') as writer:
        if not flagged_properties_df.empty:
            flagged_properties_df.to_excel(writer, sheet_name="Flagged User Properties", index=False)
        else:
            pd.DataFrame({"Message": ["No misclassified user properties found."]})\
              .to_excel(writer, sheet_name="No Data", index=False)

    print(f"User Property Misclassification Report saved to {report_file}")
